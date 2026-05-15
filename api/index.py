import os
import io
import csv
import zipfile
from contextlib import asynccontextmanager
from typing import Any, Dict, Optional
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pymongo import MongoClient

MONGODB_URI = os.getenv("MONGODB_URI", "")
DB_NAME = "valentina_tracker"
COLLECTION_NAME = "state"
USER_ID = "valentina"

client = None
db = None
collection = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    global client, db, collection
    if MONGODB_URI:
        try:
            client = MongoClient(MONGODB_URI)
            db = client[DB_NAME]
            collection = db[COLLECTION_NAME]
            print("Connected to MongoDB!")
        except Exception as e:
            print(f"Error connecting to MongoDB: {e}")
    else:
        print("Warning: MONGODB_URI not found. API will return 500 errors if DB is accessed.")
    yield
    if client:
        client.close()


app = FastAPI(title="Valentina Tracker API", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type"],
)


@app.get("/api/state")
async def get_state():
    if collection is None:
        return JSONResponse(status_code=500, content={"message": "Database not configured. Set MONGODB_URI."})

    document = collection.find_one({"_id": USER_ID})
    if document:
        document.pop("_id", None)
        return document
    return {}


@app.post("/api/state")
async def save_state(state: Dict[str, Any]):
    if collection is None:
        return JSONResponse(status_code=500, content={"message": "Database not configured. Set MONGODB_URI."})

    try:
        doc = {k: v for k, v in state.items() if k != "_id"}
        doc["_id"] = USER_ID
        collection.replace_one({"_id": USER_ID}, doc, upsert=True)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving state: {str(e)}")


# ============= EXPORT (xlsx / csv / json) =============

KPI_FIELDS = [
    ("postulaciones", "Postulaciones formales"),
    ("entrevistas",   "Entrevistas"),
    ("conexiones",    "Conexiones LinkedIn"),
    ("freelance",     "Ingreso freelance USD"),
    ("proyectos",     "Proyectos entregados"),
    ("tarifa",        "Tarifa USD/hr"),
    ("posts",         "Posts LinkedIn (inglés)"),
    ("commits",       "Commits GitHub"),
    ("certs",         "Certificaciones publicadas"),
    ("reflexion",     "Reflexión semanal"),
    ("preguntas",     "Preguntas al coach"),
    ("plan",          "Plan próxima semana"),
]


def _within(date_str: str, from_d: Optional[str], to_d: Optional[str]) -> bool:
    """ISO 8601 date strings sort lexicographically, so we compare directly."""
    if from_d and date_str < from_d:
        return False
    if to_d and date_str > to_d:
        return False
    return True


def _filter_state(state: dict, from_d: Optional[str], to_d: Optional[str]) -> dict:
    if not from_d and not to_d:
        return state
    out = dict(state)
    if isinstance(state.get("dias"), dict):
        out["dias"] = {k: v for k, v in state["dias"].items() if _within(k, from_d, to_d)}
    if isinstance(state.get("semanas"), dict):
        out["semanas"] = {k: v for k, v in state["semanas"].items() if _within(k, from_d, to_d)}
    return out


def _categoria_lookup(state: dict) -> Dict[str, dict]:
    cats = (state.get("plan") or {}).get("categorias") or []
    return {c.get("id"): c for c in cats if isinstance(c, dict)}


def _build_tables(state: dict) -> Dict[str, list]:
    """Each table is a list of dicts (long format) — first item defines the columns."""
    cat_lookup = _categoria_lookup(state)

    # 1. dias — una fila por (fecha × bloque). Wins/blockers se agregan en otra hoja para no inflar.
    dias_rows = []
    notas_rows = []
    for fecha in sorted((state.get("dias") or {}).keys()):
        d = state["dias"][fecha]
        bloques = (d.get("bloques") or {})
        for bid in sorted(bloques.keys()):
            b = bloques[bid] or {}
            cat = cat_lookup.get(bid) or {}
            dias_rows.append({
                "fecha": fecha,
                "bloque_id": bid,
                "bloque_nombre": cat.get("nombre", bid),
                "estado_categoria": cat.get("estado", ""),
                "done": bool(b.get("done")),
                "horas": float(b.get("hours") or 0),
                "que_hice": b.get("what") or "",
            })
        if d.get("wins") or d.get("blockers"):
            notas_rows.append({
                "fecha": fecha,
                "wins": d.get("wins") or "",
                "blockers": d.get("blockers") or "",
            })

    # 2. semanas — una fila por (lunes × kpi)
    semanas_rows = []
    for lunes in sorted((state.get("semanas") or {}).keys()):
        s = state["semanas"][lunes] or {}
        for kpi_id, kpi_label in KPI_FIELDS:
            v = s.get(kpi_id)
            if v in (None, ""):
                continue
            semanas_rows.append({
                "semana_inicio": lunes,
                "kpi_id": kpi_id,
                "kpi_nombre": kpi_label,
                "valor": v,
            })

    # 3. categorias
    categorias_rows = []
    for c in (state.get("plan") or {}).get("categorias") or []:
        if not isinstance(c, dict):
            continue
        categorias_rows.append({
            "id": c.get("id"),
            "nombre": c.get("nombre"),
            "meta_min": c.get("meta_min"),
            "meta_max": c.get("meta_max"),
            "color": c.get("color"),
            "orden": c.get("orden"),
            "estado": c.get("estado"),
            "creado_en": c.get("creado_en"),
            "archivado_en": c.get("archivado_en"),
        })

    # 4. libros
    libros_rows = []
    for l in state.get("libros") or []:
        if not isinstance(l, dict):
            continue
        libros_rows.append({
            "id": l.get("id"),
            "titulo": l.get("titulo") or l.get("name"),
            "autor": l.get("autor"),
            "paginas_total": l.get("paginas_total"),
            "en_plan_original": l.get("en_plan_original"),
            "estado": l.get("estado"),
            "pagina_actual": l.get("pagina_actual"),
            "inicio": l.get("inicio"),
            "fin": l.get("fin"),
            "motivo_abandono": l.get("motivo_abandono"),
            "applied_insight": l.get("applied_insight") or "",
            "horas_total": l.get("horas_total"),
            "rating": l.get("rating"),
            "orden": l.get("orden"),
        })

    # 5. certs
    certs_rows = []
    for c in state.get("certs") or []:
        if not isinstance(c, dict):
            continue
        certs_rows.append({
            "id": c.get("id"),
            "name": c.get("name"),
            "target": c.get("target"),
            "done": c.get("done"),
            "doneDate": c.get("doneDate"),
            "url_curso": c.get("url"),
            "url_certificado": c.get("certUrl"),
        })

    # 6. repos
    repos_rows = []
    for r in state.get("repos") or []:
        if not isinstance(r, dict):
            continue
        repos_rows.append({
            "id": r.get("id"),
            "name": r.get("name"),
            "target": r.get("target"),
            "done": r.get("done"),
            "doneDate": r.get("doneDate"),
            "url_repo": r.get("repoUrl"),
        })

    # 7. idiomas (array de v2; soporta dict legacy v1 también)
    idiomas_rows = []
    idiomas_data = state.get("idiomas")
    if isinstance(idiomas_data, list):
        for it in idiomas_data:
            if not isinstance(it, dict):
                continue
            idiomas_rows.append({
                "id": it.get("id"),
                "nombre": it.get("nombre"),
                "nivel": it.get("nivel"),
                "examen_nombre": it.get("examen_nombre"),
                "examen_score": it.get("examen_score"),
                "orden": it.get("orden"),
            })
    elif isinstance(idiomas_data, dict):
        # legacy: { ingles: 'B2', frances: 'A0', ielts: '...', tef: '...' }
        idiomas_rows = [
            {"id": "ingles",  "nombre": "Inglés",  "nivel": idiomas_data.get("ingles"),  "examen_nombre": "IELTS", "examen_score": idiomas_data.get("ielts"), "orden": 0},
            {"id": "frances", "nombre": "Francés", "nivel": idiomas_data.get("frances"), "examen_nombre": "TEF",   "examen_score": idiomas_data.get("tef"),   "orden": 1},
        ]

    # 8. meses (definición de cada mes del plan: id, nombre, rango de fechas)
    meses_rows = []
    for m in state.get("meses") or []:
        if not isinstance(m, dict):
            continue
        meses_rows.append({
            "id": m.get("id"),
            "nombre": m.get("nombre"),
            "desde": m.get("desde"),
            "hasta": m.get("hasta"),
            "orden": m.get("orden"),
        })

    # 9. metas_mensuales (long: una fila por mes × métrica)
    metas_rows = []
    metas_data = state.get("metas_mensuales") or {}
    mes_lookup = {m.get("id"): m for m in (state.get("meses") or []) if isinstance(m, dict)}
    if isinstance(metas_data, dict):
        for mes_id in sorted(metas_data.keys()):
            metas = metas_data[mes_id] or {}
            if not isinstance(metas, dict):
                continue
            mes_info = mes_lookup.get(mes_id) or {}
            for metrica in sorted(metas.keys()):
                metas_rows.append({
                    "mes_id": mes_id,
                    "mes_nombre": mes_info.get("nombre", mes_id),
                    "metrica": metrica,
                    "meta": metas[metrica],
                })

    # 10. tech_stack (capas L0→Ln)
    tech_rows = []
    for l in state.get("tech_stack_layers") or []:
        if not isinstance(l, dict):
            continue
        tech_rows.append({
            "id": l.get("id"),
            "orden": l.get("orden"),
            "codigo": l.get("codigo"),
            "nombre": l.get("nombre"),
            "estado": l.get("estado"),
            "evidencia_url": l.get("evidencia_url"),
            "notas": l.get("notas"),
            "fecha_completado": l.get("fecha_completado"),
        })

    # 11. rutas_carrera (resumen — una fila por ruta con conteo de milestones)
    # 12. rutas_milestones (detalle long — una fila por (ruta × milestone))
    rutas_rows = []
    rutas_milestones_rows = []
    for r in state.get("rutas_carrera") or []:
        if not isinstance(r, dict):
            continue
        milestones = r.get("milestones") or []
        total = sum(1 for m in milestones if isinstance(m, dict))
        done = sum(1 for m in milestones if isinstance(m, dict) and m.get("completado"))
        pct = round((done / total) * 100) if total else 0
        rutas_rows.append({
            "id": r.get("id"),
            "nombre": r.get("nombre"),
            "emoji": r.get("emoji"),
            "estado": r.get("estado"),
            "fecha_inicio": r.get("fecha_inicio"),
            "descripcion": r.get("descripcion"),
            "analisis_estrategico": r.get("analisis"),
            "milestones_total": total,
            "milestones_completados": done,
            "pct": pct,
        })
        for m in milestones:
            if not isinstance(m, dict):
                continue
            rutas_milestones_rows.append({
                "ruta_id": r.get("id"),
                "ruta_nombre": r.get("nombre"),
                "milestone_id": m.get("id"),
                "milestone_texto": m.get("texto"),
                "completado": bool(m.get("completado")),
                "fecha_completado": m.get("fecha_completado"),
                "notas": m.get("notas") or "",
            })

    # 13. decisiones_historial (filtro 6 verdades)
    decisiones_rows = []
    df = state.get("decision_filter") or {}
    for h in df.get("historial") or []:
        if not isinstance(h, dict):
            continue
        decisiones_rows.append({
            "fecha": h.get("fecha"),
            "decision": h.get("decision"),
            "verdades_aprobadas": h.get("verdades_aprobadas"),
            "umbral": h.get("umbral"),
            "resultado": h.get("resultado"),
        })

    # 14. meta plan (resumen ejecutivo enriquecido con las nuevas features)
    plan = state.get("plan") or {}
    artefacto_settings = state.get("artefacto_settings") or {}
    freelance_tier = state.get("freelance_tier") or {}
    semanas = state.get("semanas") or {}
    resenas_total = sum(int((s or {}).get("resenas") or 0) for s in semanas.values()) if isinstance(semanas, dict) else 0
    meta_rows = [
        {"campo": "schema_version", "valor": state.get("schema_version", "")},
        {"campo": "plan_nombre", "valor": plan.get("nombre", "")},
        {"campo": "categorias_activas", "valor": sum(1 for c in (plan.get("categorias") or []) if c.get("estado") == "activo")},
        {"campo": "categorias_archivadas", "valor": sum(1 for c in (plan.get("categorias") or []) if c.get("estado") == "archivado")},
        {"campo": "dias_registrados", "valor": len(state.get("dias") or {})},
        {"campo": "semanas_kpi", "valor": len(semanas) if isinstance(semanas, dict) else 0},
        {"campo": "libros_totales", "valor": len(state.get("libros") or [])},
        {"campo": "libros_terminados", "valor": sum(1 for l in (state.get("libros") or []) if l.get("estado") == "terminado")},
        {"campo": "meses_definidos", "valor": len(state.get("meses") or [])},
        {"campo": "tech_stack_total", "valor": len(state.get("tech_stack_layers") or [])},
        {"campo": "tech_stack_completados", "valor": sum(1 for l in (state.get("tech_stack_layers") or []) if l.get("estado") == "completado")},
        {"campo": "rutas_activas", "valor": sum(1 for r in (state.get("rutas_carrera") or []) if r.get("estado") == "activa")},
        {"campo": "rutas_pausadas", "valor": sum(1 for r in (state.get("rutas_carrera") or []) if r.get("estado") == "pausada")},
        {"campo": "rutas_descartadas", "valor": sum(1 for r in (state.get("rutas_carrera") or []) if r.get("estado") == "descartada")},
        {"campo": "artefacto_categorias_obligatorias", "valor": ", ".join(artefacto_settings.get("categorias_obligatorias") or [])},
        {"campo": "freelance_tier_actual", "valor": freelance_tier.get("ultimo_tier_visto", "")},
        {"campo": "freelance_resenas_acumuladas", "valor": resenas_total},
        {"campo": "freelance_plataformas", "valor": ", ".join(freelance_tier.get("plataformas") or [])},
        {"campo": "filtro_decision_umbral", "valor": df.get("umbral", "")},
        {"campo": "filtro_decision_verdades_count", "valor": len(df.get("verdades") or [])},
        {"campo": "filtro_decision_historial_count", "valor": len(df.get("historial") or [])},
    ]

    return {
        "dias": dias_rows,
        "semanas_kpi": semanas_rows,
        "categorias": categorias_rows,
        "libros": libros_rows,
        "certificaciones": certs_rows,
        "repos_github": repos_rows,
        "idiomas": idiomas_rows,
        "notas_diarias": notas_rows,
        "meses": meses_rows,
        "metas_mensuales": metas_rows,
        "tech_stack": tech_rows,
        "rutas_carrera": rutas_rows,
        "rutas_milestones": rutas_milestones_rows,
        "decisiones_historial": decisiones_rows,
        "meta_plan": meta_rows,
    }


def _build_csv_zip(state: dict) -> bytes:
    tables = _build_tables(state)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, rows in tables.items():
            sbuf = io.StringIO()
            if rows:
                fieldnames = list(rows[0].keys())
                w = csv.DictWriter(sbuf, fieldnames=fieldnames, extrasaction="ignore")
                w.writeheader()
                for r in rows:
                    w.writerow(r)
            else:
                sbuf.write("(sin registros)\n")
            zf.writestr(f"{name}.csv", sbuf.getvalue())
    return buf.getvalue()


def _build_xlsx(state: dict) -> bytes:
    # Import diferido para arranque más rápido cuando el endpoint no se usa
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    tables = _build_tables(state)
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="1F4E79")
    header_fill = PatternFill("solid", fgColor="D9E2F3")

    for name, rows in tables.items():
        ws = wb.create_sheet(title=name[:31])  # límite Excel
        if not rows:
            ws.append(["(sin registros)"])
            continue
        fieldnames = list(rows[0].keys())
        ws.append(fieldnames)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
        for r in rows:
            ws.append([r.get(k) for k in fieldnames])
        # Ancho aproximado por columna (cap a 40)
        for i, col in enumerate(fieldnames, 1):
            max_len = max([len(str(col))] + [len(str(r.get(col, ""))) for r in rows[:200]])
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(40, max(10, max_len + 2))
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/api/export")
async def export_data(
    format: str = Query("json", regex="^(json|csv|xlsx)$"),
    from_date: Optional[str] = Query(None, alias="from", regex=r"^\d{4}-\d{2}-\d{2}$"),
    to_date: Optional[str] = Query(None, alias="to", regex=r"^\d{4}-\d{2}-\d{2}$"),
):
    if collection is None:
        return JSONResponse(status_code=500, content={"message": "Database not configured."})

    doc = collection.find_one({"_id": USER_ID}) or {}
    doc.pop("_id", None)
    filtered = _filter_state(doc, from_date, to_date)

    range_part = ""
    if from_date or to_date:
        range_part = f"_{from_date or 'inicio'}_a_{to_date or 'fin'}"

    if format == "json":
        return JSONResponse(
            filtered,
            headers={"Content-Disposition": f'attachment; filename="valentina{range_part}.json"'},
        )
    if format == "csv":
        data = _build_csv_zip(filtered)
        return Response(
            data,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="valentina{range_part}_csv.zip"'},
        )
    if format == "xlsx":
        data = _build_xlsx(filtered)
        return Response(
            data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="valentina{range_part}.xlsx"'},
        )

    raise HTTPException(status_code=400, detail="format must be json|csv|xlsx")
